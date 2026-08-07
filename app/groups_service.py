"""
groups_service.py — ตัวกลางสำหรับ group-mail audit (หลาย workspace)
- อ่าน config workspace จาก env WORKSPACES (JSON) หรือ fallback เป็น workspace เดียว
- รัน crawl ทุก workspace แล้วเขียน groups.json / groups.xlsx
"""

import os
import json
import threading
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from groups_audit import audit_workspace

# ---- สถานะ crawl (in-memory) ----
GROUPS_STATE = {
    "running": False,
    "started": None,
    "last_run": None,
    "last_error": None,
    "phase": "",          # ข้อความบอกความคืบหน้า
    "workspaces": 0,
    "groups": 0,
}
_lock = threading.Lock()


def load_workspaces():
    """
    คืน list ของ workspace: [{"id","name","sa_file","admin","domains"(set)}]

    วิธีตั้งค่า (เลือกอย่างใดอย่างหนึ่ง):
    1) env WORKSPACES = JSON list เช่น
       [{"id":"office21sun","name":"office21sun.com",
         "admin":"admin@office21sun.com",
         "domains":"office21sun.com,21sunpassion.com,sunsusolution.com"},
        {"id":"21sunpassion","name":"21sunpassion.com",
         "admin":"admin@21sunpassion.com",
         "domains":"21sunpassion.com",
         "sa_file":"/secrets/sa-21sun.json"}]
       - "sa_file" ใส่หรือไม่ก็ได้ ถ้าไม่ใส่จะใช้ค่า default จาก env SA_FILE
         (กรณี SA ตัวเดียวได้ DWD ทั้ง 2 tenant)
    2) ถ้าไม่ตั้ง WORKSPACES จะ fallback ใช้ ADMIN_EMAIL / INTERNAL_DOMAINS / SA_FILE
       เป็น workspace เดียว (เข้ากับ config เดิม)
    """
    default_sa = os.getenv("SA_FILE", "sa.json")
    raw = os.getenv("WORKSPACES", "").strip()

    if raw:
        items = json.loads(raw)
    else:
        items = [{
            "id": "default",
            "name": os.getenv("ADMIN_EMAIL", "default"),
            "admin": os.getenv("ADMIN_EMAIL", ""),
            "domains": os.getenv(
                "INTERNAL_DOMAINS",
                "office21sun.com,21sunpassion.com,sunsusolution.com"),
        }]

    out = []
    for it in items:
        dom = it.get("domains", "")
        if isinstance(dom, str):
            dom = {d.strip().lower() for d in dom.split(",") if d.strip()}
        else:
            dom = {str(d).strip().lower() for d in dom if str(d).strip()}
        out.append({
            "id": str(it.get("id") or it.get("name") or it.get("admin")),
            "name": str(it.get("name") or it.get("admin") or it.get("id")),
            "sa_file": it.get("sa_file") or default_sa,
            "admin": it.get("admin", ""),
            "domains": dom,
        })
    return out


# -------------------------------------------------- Excel
HEADER_FILL = PatternFill("solid", fgColor="404040")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws, max_w=70):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None),
                    default=10)
        ws.column_dimensions[letter].width = min(width + 2, max_w)


def write_xlsx(data, path):
    wb = Workbook()

    # ---- Sheet 1: Groups ----
    ws = wb.active
    ws.title = "Groups"
    ws.append(["Workspace", "Group Email", "Name", "Members", "External",
               "Owners", "Managers", "Allow External", "Description"])
    for w in data["workspaces"]:
        for g in w["groups"]:
            ws.append([
                w["name"], g["email"], g["name"], g["memberCount"],
                g["externalCount"], g["roleCount"].get("OWNER", 0),
                g["roleCount"].get("MANAGER", 0),
                "yes" if g["allowExternalMembers"] else "",
                g["description"],
            ])
    _style(ws, 9)
    _autofit(ws)

    # ---- Sheet 2: Members ----
    wm = wb.create_sheet("Members")
    wm.append(["Workspace", "Group Email", "Member", "Role", "Type",
               "Status", "External"])
    for w in data["workspaces"]:
        for g in w["groups"]:
            for m in g["members"]:
                wm.append([w["name"], g["email"], m["email"], m["role"],
                           m["type"], m["status"],
                           "yes" if m["external"] else ""])
    _style(wm, 7)
    _autofit(wm)

    wb.save(path)


# -------------------------------------------------- run
def run_groups_audit(out_dir):
    """crawl group-mail ทุก workspace (thread-safe)"""
    with _lock:
        if GROUPS_STATE["running"]:
            return {"skipped": True, "reason": "already running"}
        GROUPS_STATE["running"] = True
        GROUPS_STATE["started"] = dt.datetime.now().isoformat(timespec="seconds")
        GROUPS_STATE["last_error"] = None
        GROUPS_STATE["phase"] = "เริ่ม…"

    try:
        wss = load_workspaces()
        wss = [w for w in wss if w.get("admin")]
        if not wss:
            raise RuntimeError("ยังไม่ตั้ง workspace/admin ใน environment")

        result = {"workspaces": [],
                  "generated": None}
        total_groups = 0
        for w in wss:
            GROUPS_STATE["phase"] = f"กำลังดึง {w['name']}…"

            def _p(i, tot, _w=w):
                GROUPS_STATE["phase"] = f"{_w['name']}: {i}/{tot} กลุ่ม"

            wr = audit_workspace(w, out_progress=_p)
            result["workspaces"].append(wr)
            total_groups += wr["groupCount"]

        result["generated"] = dt.datetime.now().isoformat(timespec="seconds")

        os.makedirs(out_dir, exist_ok=True)
        with open(os.path.join(out_dir, "groups.json"), "w",
                  encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=1)
        write_xlsx(result, os.path.join(out_dir, "groups.xlsx"))

        GROUPS_STATE.update(workspaces=len(wss), groups=total_groups,
                            last_run=result["generated"], phase="")
        return {"workspaces": len(wss), "groups": total_groups}
    except Exception as e:  # noqa
        GROUPS_STATE["last_error"] = str(e)
        GROUPS_STATE["phase"] = ""
        raise
    finally:
        GROUPS_STATE["running"] = False
