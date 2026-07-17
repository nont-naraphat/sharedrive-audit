"""
audit_service.py — ตัวกลางที่ทั้ง CLI และ API เรียกใช้
"""

import os
import json
import csv
import threading
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from gdrive_audit import audit_all, iter_permission_rows

# ---- สถานะ crawl (in-memory) ----
STATE = {
    "running": False,
    "started": None,
    "last_run": None,
    "last_error": None,
    "drives": 0,
    "files": 0,
    "folders": 0,
}
_lock = threading.Lock()

# ---- สถานะ export permission CSV ----
PERM_STATE = {
    "running": False,
    "rows": 0,
    "last_run": None,
    "last_error": None,
}
_perm_lock = threading.Lock()


def human_size(n):
    if not n:
        return ""
    units = ["B", "KB", "MB", "GB", "TB"]
    f = float(n)
    for u in units:
        if f < 1024 or u == units[-1]:
            return f"{int(f)} B" if u == "B" else f"{f:.2f} {u}"
        f /= 1024


def _flatten(node, drive_name, rows):
    for c in (node.get("children") or []):
        rows.append({
            "drive": drive_name, "path": c["path"], "type": c["type"],
            "name": c["name"], "mimeType": c.get("mimeType"), "size": c.get("size"),
            "created": c.get("createdTime"), "modified": c.get("modifiedTime"),
            "lastModifiedBy": c.get("lastModifiedBy"), "link": c.get("webViewLink"),
        })
        if c["type"] == "folder":
            _flatten(c, drive_name, rows)


HEADER_FILL = PatternFill("solid", fgColor="404040")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws, max_w=70):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(width + 2, max_w)


def write_xlsx(data, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Drive", "Files", "Folders", "Size", "Members", "Created"])
    for s in data["summary"]:
        ws.append([s["drive"], s["files"], s["folders"],
                   human_size(s["totalSize"]), s["members"], s["createdTime"]])
    _style(ws, 6)
    _autofit(ws)

    wf = wb.create_sheet("Files")
    wf.append(["Drive", "Path", "Type", "Name", "MimeType", "Size",
               "Created", "Modified", "Last Modified By", "Link"])
    for root in data["drives"]:
        rows = []
        _flatten(root, root["name"], rows)
        for r in rows:
            wf.append([r["drive"], r["path"], r["type"], r["name"], r["mimeType"],
                       human_size(r["size"]), r["created"], r["modified"],
                       r["lastModifiedBy"], r["link"]])
    _style(wf, 10)
    _autofit(wf)

    wm = wb.create_sheet("Members")
    wm.append(["Drive", "Member", "Display Name", "Role", "Type", "Domain"])
    id2name = {s["driveId"]: s["drive"] for s in data["summary"]}
    for drive_id, members in data["members"].items():
        for m in members:
            if m.get("deleted"):
                continue
            wm.append([id2name.get(drive_id, drive_id), m.get("emailAddress", ""),
                       m.get("displayName", ""), m.get("role", ""),
                       m.get("type", ""), m.get("domain", "")])
    _style(wm, 6)
    _autofit(wm)
    wb.save(path)


def run_audit(sa_file, admin_email, internal_domains, out_dir):
    """รัน crawl ทั้งโดเมน (thread-safe: กันรันซ้อน)"""
    with _lock:
        if STATE["running"]:
            return {"skipped": True, "reason": "already running"}
        STATE["running"] = True
        STATE["started"] = dt.datetime.now().isoformat(timespec="seconds")
        STATE["last_error"] = None

    try:
        data = audit_all(sa_file, admin_email, internal_domains)
        data["generated"] = dt.datetime.now().isoformat(timespec="seconds")

        os.makedirs(out_dir, exist_ok=True)
        with open(os.path.join(out_dir, "audit.json"), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
        write_xlsx(data, os.path.join(out_dir, "audit.xlsx"))

        files = sum(s["files"] for s in data["summary"])
        folders = sum(s["folders"] for s in data["summary"])
        STATE.update(drives=len(data["drives"]), files=files, folders=folders,
                     last_run=data["generated"])
        return {"drives": len(data["drives"]), "files": files, "folders": folders}
    except Exception as e:  # noqa
        STATE["last_error"] = str(e)
        raise
    finally:
        STATE["running"] = False


PERM_CSV_HEADER = ["Drive", "Path", "Type", "Name", "Item ID", "Member",
                   "Member Type", "Role", "Role(TH)", "Inherited",
                   "External", "Domain"]


def export_permissions_csv(sa_file, admin_email, internal_domains, out_dir):
    """ไล่ทุกไฟล์ทั้งระบบ ดึง permission แล้วเขียน permissions.csv (stream ทีละแถว)"""
    with _perm_lock:
        if PERM_STATE["running"]:
            return {"skipped": True}
        PERM_STATE["running"] = True
        PERM_STATE["rows"] = 0
        PERM_STATE["last_error"] = None

    os.makedirs(out_dir, exist_ok=True)
    tmp = os.path.join(out_dir, "permissions.csv.tmp")
    final = os.path.join(out_dir, "permissions.csv")

    def _progress(n):
        PERM_STATE["rows"] = n
        print(f"     [perms] เขียนแล้ว {n:,} แถว", flush=True)

    try:
        with open(tmp, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            w.writerow(PERM_CSV_HEADER)
            for row in iter_permission_rows(sa_file, admin_email,
                                            internal_domains, progress=_progress):
                w.writerow([row["drive"], row["path"], row["type"], row["name"],
                            row["item_id"], row["member"], row["member_type"],
                            row["role"], row["role_th"], row["inherited"],
                            row["external"], row["domain"]])
        os.replace(tmp, final)
        PERM_STATE["last_run"] = dt.datetime.now().isoformat(timespec="seconds")
        print(f"[✓] permissions.csv เสร็จ ({PERM_STATE['rows']:,} แถว)", flush=True)
        return {"rows": PERM_STATE["rows"]}
    except Exception as e:  # noqa
        PERM_STATE["last_error"] = str(e)
        raise
    finally:
        PERM_STATE["running"] = False
