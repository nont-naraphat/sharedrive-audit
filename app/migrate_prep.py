"""
migrate_prep.py — ชั้นเตรียม migration (Google Shared Drive → SharePoint / M365)

ทำงานจาก output ที่ audit tool เดิมสร้างไว้แล้ว:
  - audit.json       (โครงสร้างไฟล์/โฟลเดอร์ + members ต่อ drive)
  - permissions.csv  (optional — permission ราย item สำหรับ external report)

สร้าง artifact สำหรับ Microsoft Migration Manager + งาน planning:
  migrate/
    mm_source_paths.csv     -> อัปโหลดตรงใน Migration Manager (bulk source, /ชื่อ drive)
    mm_identity_map.csv      -> ตารางเทียบ Google identity -> M365 UPN (auto-guess)
    destination_plan.csv     -> ตารางวางแผน drive -> SharePoint site/library (แก้เองได้)
    readiness_report.xlsx    -> รายงานสิ่งที่ต้องแก้ก่อนย้าย (blocker scan)

หมายเหตุสำคัญ (ยืนยันจากเอกสาร Microsoft ปี 2026):
  - Migration Manager คือ tool ทางการที่ย้าย Google Drive/Shared Drive (Mover ถูก retire แล้ว)
  - permission ของ Shared Drive แทบไม่ถูกย้าย -> ต้องใช้ readiness/permission report ช่วย
  - path (URL) เกิน 400 ตัวจะย้ายไม่ได้, ไฟล์ใหญ่เกิน 250 GB ไม่รองรับ
  - Google native (doc/sheet/slide) จะถูกแปลงอัตโนมัติ; drawing/jam แปลงแบบแก้ไม่ได้;
    form/site/map/script ไม่ถูกย้าย
"""

import os
import csv
import json
import re
import threading
import datetime as dt

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------------------------------------------------------- state
PREP_STATE = {
    "running": False,
    "last_run": None,
    "last_error": None,
    "drives": 0,
    "items": 0,
    "blockers": 0,
    "warnings": 0,
    "summary": {},          # นับ issue แต่ละชนิด สำหรับโชว์บนหน้าเว็บ
}
_lock = threading.Lock()

# ---------------------------------------------------------------- constants
GOOGLE_NATIVE = {
    "application/vnd.google-apps.document":     ("Google Docs",        "แปลงเป็น .docx",   "ok"),
    "application/vnd.google-apps.spreadsheet":  ("Google Sheets",      "แปลงเป็น .xlsx",   "ok"),
    "application/vnd.google-apps.presentation": ("Google Slides",      "แปลงเป็น .pptx",   "ok"),
    "application/vnd.google-apps.drawing":      ("Google Drawing",     "แปลงเป็น .jpg (แก้ไขต่อไม่ได้)", "warn"),
    "application/vnd.google-apps.jam":          ("Jamboard",           "แปลงเป็น .pdf (แก้ไขต่อไม่ได้)", "warn"),
    "application/vnd.google-apps.form":         ("Google Form",        "ไม่ถูกย้าย",       "block"),
    "application/vnd.google-apps.site":         ("Google Site",        "ไม่ถูกย้าย",       "block"),
    "application/vnd.google-apps.map":          ("Google My Maps",     "ไม่ถูกย้าย",       "block"),
    "application/vnd.google-apps.script":       ("Apps Script",        "ไม่ถูกย้าย",       "block"),
    "application/vnd.google-apps.shortcut":     ("Shortcut",           "shortcut ไม่ถูกย้าย (เป็น pointer)", "warn"),
    "application/vnd.google-apps.fusiontable":  ("Fusion Table",       "ไม่ถูกย้าย",       "block"),
}

# อักขระที่ SharePoint/OneDrive ไม่รับในชื่อไฟล์/โฟลเดอร์
ILLEGAL_CHARS = set('"*:<>?/\\|')
# ชื่อสงวน
RESERVED_NAMES = {
    "con", "prn", "aux", "nul", "desktop.ini", ".lock",
    *(f"com{i}" for i in range(0, 10)),
    *(f"lpt{i}" for i in range(0, 10)),
}

MAX_PATH = 400            # ความยาว path (URL) สูงสุดฝั่ง SharePoint
PATH_WARN = 350           # เตือนล่วงหน้าก่อนถึง limit
MAX_FILE_BYTES = 250 * 1024 ** 3   # 250 GB


def _slug_site(name: str) -> str:
    """ทำชื่อ drive ให้พอเป็น SharePoint site name / URL segment ได้"""
    s = name.strip()
    s = re.sub(r'[\\/:*?"<>|#%&{}+~]', "", s)   # ตัดเฉพาะอักขระที่ URL/SharePoint ไม่รับ (คงภาษาไทยไว้ครบ)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "site"


def _name_issues(name: str):
    """คืน list ปัญหาของ 'ชื่อ' item ตามกฎ SharePoint"""
    issues = []
    bad = sorted({c for c in name if c in ILLEGAL_CHARS})
    if bad:
        issues.append(("ILLEGAL_CHARS", "block", "ชื่อมีอักขระต้องห้าม: " + " ".join(bad)))
    if name != name.strip():
        issues.append(("EDGE_SPACE", "warn", "ชื่อมีเว้นวรรคหน้า/หลัง"))
    if name.endswith("."):
        issues.append(("TRAILING_DOT", "block", "ชื่อลงท้ายด้วยจุด"))
    if name.startswith("~$"):
        issues.append(("TEMP_PREFIX", "warn", "ชื่อขึ้นต้น ~$ (มักเป็นไฟล์ temp)"))
    if "_vti_" in name.lower():
        issues.append(("VTI", "block", "ชื่อมี _vti_ (สงวนโดย SharePoint)"))
    stem = name.split(".")[0].lower() if "." in name else name.lower()
    if name.lower() in RESERVED_NAMES or stem in RESERVED_NAMES:
        issues.append(("RESERVED_NAME", "block", "เป็นชื่อสงวนของ Windows/SharePoint"))
    return issues


# ---------------------------------------------------------------- walk
def _walk_items(node, drive_name):
    """เดินต้นไม้ audit.json คืน dict ราย item (path, name, mimeType, size ...)"""
    for c in (node.get("children") or []):
        yield {
            "drive": drive_name,
            "path": c.get("path") or (drive_name + "/" + c.get("name", "")),
            "name": c.get("name", ""),
            "type": c.get("type"),
            "mimeType": c.get("mimeType"),
            "size": c.get("size"),
        }
        if c.get("type") == "folder":
            yield from _walk_items(c, drive_name)


def scan_items(audit):
    """สแกนทุก item หา blocker/warning; คืน (rows, native_rows, counter)"""
    rows = []          # แถวที่ต้องแก้ (มี issue)
    native_rows = []   # inventory ของ Google native ทั้งหมด
    counter = {}

    def bump(code):
        counter[code] = counter.get(code, 0) + 1

    for root in audit.get("drives", []):
        dname = root.get("name", "")
        for it in _walk_items(root, dname):
            name = it["name"]
            path = it["path"]
            mime = it.get("mimeType") or ""

            item_issues = list(_name_issues(name))

            plen = len(path)
            if plen > MAX_PATH:
                item_issues.append(("LONG_PATH", "block", f"path ยาว {plen} ตัว (>{MAX_PATH})"))
            elif plen > PATH_WARN:
                item_issues.append(("PATH_NEAR_LIMIT", "warn", f"path ยาว {plen} ตัว (ใกล้ลิมิต {MAX_PATH})"))

            size = it.get("size")
            if isinstance(size, (int, float)) and size > MAX_FILE_BYTES:
                item_issues.append(("OVERSIZE", "block", "ไฟล์ใหญ่เกิน 250 GB"))

            if mime in GOOGLE_NATIVE:
                label, note, sev = GOOGLE_NATIVE[mime]
                native_rows.append({
                    "drive": dname, "path": path, "name": name,
                    "kind": label, "result": note, "severity": sev,
                })
                bump("native_" + sev)
                if sev in ("block", "warn"):
                    item_issues.append(("NATIVE_" + sev.upper(), sev,
                                        f"{label}: {note}"))

            for code, sev, msg in item_issues:
                bump(sev)
                bump(code)
                rows.append({
                    "drive": dname, "path": path, "name": name,
                    "type": it["type"], "code": code,
                    "severity": sev, "detail": msg,
                })
    return rows, native_rows, counter


# ---------------------------------------------------------------- identities
def collect_identities(audit, perms_csv_path, internal_domains):
    """รวม email ทั้งหมดที่พบ (จาก members + permissions.csv) -> เดา M365 UPN"""
    seen = {}   # email -> {"display", "internal", "sources": set}

    def add(email, display, source):
        if not email or "@" not in email:
            return
        e = email.strip().lower()
        rec = seen.setdefault(e, {"display": display or "", "sources": set()})
        rec["sources"].add(source)
        if display and not rec["display"]:
            rec["display"] = display

    for members in (audit.get("members") or {}).values():
        for m in members:
            if m.get("deleted"):
                continue
            add(m.get("emailAddress"), m.get("displayName"), "drive-member")

    if perms_csv_path and os.path.exists(perms_csv_path):
        with open(perms_csv_path, encoding="utf-8-sig", newline="") as fh:
            for r in csv.DictReader(fh):
                add(r.get("Member"), "", "item-permission")

    rows = []
    for email, rec in sorted(seen.items()):
        dom = email.split("@")[-1]
        internal = dom in internal_domains
        # เดา UPN: ภายในองค์กร -> ใช้ email เดิม (Migration Manager automap แบบ exact match)
        guess = email if internal else ""
        rows.append({
            "source": email,
            "m365": guess,
            "display": rec["display"],
            "internal": "yes" if internal else "no",
            "status": "auto" if internal else "REVIEW (external/guest)",
            "seen_in": ", ".join(sorted(rec["sources"])),
        })
    return rows


def collect_external_shares(perms_csv_path):
    """ดึงเฉพาะ permission ที่ external=yes จาก permissions.csv (ถ้ามี)"""
    out = []
    if not (perms_csv_path and os.path.exists(perms_csv_path)):
        return out
    with open(perms_csv_path, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            if (r.get("External") or "").lower() == "yes":
                out.append(r)
    return out


# ---------------------------------------------------------------- xlsx (streaming)
_HEAD_FILL = "404040"
_SEV_FILL = {"block": "F8D7DA", "warn": "FFF3CD", "ok": "D4EDDA"}


def _ws_write_only(wb, title, header):
    ws = wb.create_sheet(title)
    hdr = []
    for h in header:
        c = WriteOnlyCell(ws, value=h)
        c.fill = PatternFill("solid", fgColor=_HEAD_FILL)
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(vertical="center")
        hdr.append(c)
    ws.append(hdr)
    ws.freeze_panes = "A2"
    return ws


def _sev_cell(ws, value, sev):
    c = WriteOnlyCell(ws, value=value)
    fill = _SEV_FILL.get(sev)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = Font(bold=(sev == "block"))
    return c


def write_readiness_xlsx(path, counter, rows, native_rows, ext_rows, drives_meta):
    """เขียนรายงาน readiness แบบ write_only (กัน OOM ตามหลักการที่ใช้ประจำ)"""
    wb = Workbook(write_only=True)

    # ---- Summary
    ws = _ws_write_only(wb, "Summary", ["หัวข้อ", "จำนวน"])
    order = [
        ("รวม item ที่ต้องแก้ (block)",        counter.get("block", 0)),
        ("รวม item ที่ต้องระวัง (warn)",       counter.get("warn", 0)),
        ("path ยาวเกิน 400",                   counter.get("LONG_PATH", 0)),
        ("path ใกล้ลิมิต (>350)",              counter.get("PATH_NEAR_LIMIT", 0)),
        ("ชื่อมีอักขระต้องห้าม",              counter.get("ILLEGAL_CHARS", 0)),
        ("ชื่อสงวน",                           counter.get("RESERVED_NAME", 0)),
        ("เว้นวรรคหน้า/หลังชื่อ",             counter.get("EDGE_SPACE", 0)),
        ("ไฟล์ใหญ่เกิน 250 GB",                counter.get("OVERSIZE", 0)),
        ("Google native ที่ไม่ถูกย้าย",        counter.get("native_block", 0)),
        ("Google native ที่แปลงแบบ lossy",     counter.get("native_warn", 0)),
        ("Google native ที่แปลงได้ปกติ",       counter.get("native_ok", 0)),
        ("External share (item-level)",         len(ext_rows)),
    ]
    for label, n in order:
        ws.append([label, n])

    # ---- Blockers / Warnings
    ws = _ws_write_only(wb, "Action Needed",
                        ["Drive", "Path", "Name", "Type", "Severity", "Issue", "Detail"])
    for r in sorted(rows, key=lambda x: (x["severity"] != "block", x["drive"])):
        ws.append([
            r["drive"], r["path"], r["name"], r["type"],
            _sev_cell(ws, r["severity"], r["severity"]),
            r["code"], r["detail"],
        ])

    # ---- Native inventory
    ws = _ws_write_only(wb, "Google Native",
                        ["Drive", "Path", "Name", "Kind", "ผลหลังย้าย", "Severity"])
    for r in native_rows:
        ws.append([
            r["drive"], r["path"], r["name"], r["kind"], r["result"],
            _sev_cell(ws, r["severity"], r["severity"]),
        ])

    # ---- External shares
    ws = _ws_write_only(wb, "External Shares",
                        ["Drive", "Path", "Name", "Member", "Role", "Domain"])
    for r in ext_rows:
        ws.append([r.get("Drive", ""), r.get("Path", ""), r.get("Name", ""),
                   r.get("Member", ""), r.get("Role", ""), r.get("Domain", "")])

    # ---- Per-drive size (ช่วยวางแผน batch)
    ws = _ws_write_only(wb, "Drive Sizing",
                        ["Drive", "Files", "Folders", "Size (GB)", "Members"])
    for s in drives_meta:
        gb = round((s.get("totalSize") or 0) / 1024 ** 3, 2)
        ws.append([s.get("drive", ""), s.get("files", 0), s.get("folders", 0),
                   gb, s.get("members", 0)])

    wb.save(path)


# ---------------------------------------------------------------- csv artifacts
def write_source_csv(path, drives_meta):
    """Migration Manager bulk source: หนึ่ง shared drive ต่อแถว รูปแบบ /ชื่อ drive"""
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        for s in drives_meta:
            w.writerow(["/" + s.get("drive", "")])


def write_identity_csv(path, id_rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["Source identity", "Microsoft 365 identity",
                    "Display name", "Internal", "Status", "Seen in"])
        for r in id_rows:
            w.writerow([r["source"], r["m365"], r["display"],
                        r["internal"], r["status"], r["seen_in"]])


def write_destination_csv(path, drives_meta, site_base, target_domain):
    """ตารางวางแผน (แก้เองได้) — 1 drive : 1 site เป็น default"""
    base = (site_base or "").rstrip("/")
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["Shared Drive", "Files", "Folders", "Size (GB)",
                    "Suggested SharePoint Site", "Suggested Library", "Notes"])
        for s in drives_meta:
            dname = s.get("drive", "")
            slug = _slug_site(dname)
            site_url = f"{base}/{slug}" if base else f".../sites/{slug}"
            gb = round((s.get("totalSize") or 0) / 1024 ** 3, 2)
            w.writerow([dname, s.get("files", 0), s.get("folders", 0), gb,
                        site_url, "Documents", ""])


# ---------------------------------------------------------------- orchestrator
def build_migration_pack(audit_json_path, out_dir, internal_domains,
                         site_base="", target_domain="", perms_csv_path=None):
    with _lock:
        if PREP_STATE["running"]:
            return {"skipped": True}
        PREP_STATE["running"] = True
        PREP_STATE["last_error"] = None

    try:
        if not os.path.exists(audit_json_path):
            raise FileNotFoundError("ยังไม่มี audit.json — กด Sync ในหน้าหลักก่อน")

        with open(audit_json_path, encoding="utf-8") as f:
            audit = json.load(f)

        drives_meta = audit.get("summary", [])
        if perms_csv_path is None:
            perms_csv_path = os.path.join(os.path.dirname(audit_json_path), "permissions.csv")

        rows, native_rows, counter = scan_items(audit)
        id_rows = collect_identities(audit, perms_csv_path, internal_domains)
        ext_rows = collect_external_shares(perms_csv_path)

        mdir = os.path.join(out_dir, "migrate")
        os.makedirs(mdir, exist_ok=True)

        write_source_csv(os.path.join(mdir, "mm_source_paths.csv"), drives_meta)
        write_identity_csv(os.path.join(mdir, "mm_identity_map.csv"), id_rows)
        write_destination_csv(os.path.join(mdir, "destination_plan.csv"),
                              drives_meta, site_base, target_domain)
        write_readiness_xlsx(os.path.join(mdir, "readiness_report.xlsx"),
                             counter, rows, native_rows, ext_rows, drives_meta)

        items = sum(1 for _ in (r for root in audit.get("drives", [])
                                for r in _walk_items(root, root.get("name", ""))))
        summary = {
            "block": counter.get("block", 0),
            "warn": counter.get("warn", 0),
            "long_path": counter.get("LONG_PATH", 0),
            "illegal": counter.get("ILLEGAL_CHARS", 0),
            "oversize": counter.get("OVERSIZE", 0),
            "native_block": counter.get("native_block", 0),
            "native_warn": counter.get("native_warn", 0),
            "native_ok": counter.get("native_ok", 0),
            "external": len(ext_rows),
            "identities": len(id_rows),
        }
        PREP_STATE.update(
            drives=len(drives_meta), items=items,
            blockers=summary["block"], warnings=summary["warn"],
            summary=summary,
            last_run=dt.datetime.now().isoformat(timespec="seconds"),
        )
        return {"drives": len(drives_meta), "items": items, **summary}
    except Exception as e:  # noqa
        PREP_STATE["last_error"] = str(e)
        raise
    finally:
        PREP_STATE["running"] = False


ARTIFACTS = {
    "source":      ("mm_source_paths.csv",   "text/csv"),
    "identity":    ("mm_identity_map.csv",   "text/csv"),
    "destination": ("destination_plan.csv",  "text/csv"),
    "readiness":   ("readiness_report.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
}
